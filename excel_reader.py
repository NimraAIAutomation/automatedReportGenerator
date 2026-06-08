import openpyxl

filepath = "data/sales_report_dummy.xlsx"
def read_excel(filepath):
    # Load the Excel workbook from the given file path
    wb = openpyxl.load_workbook(filepath)
    
    # Select the active (first) sheet
    ws = wb.active
    
    # Read the first row as column headers
    # ws[1] returns all cells in row 1
    headers = [cell.value for cell in ws[1]]
    
    rows = []
    
    # Loop through all rows starting from row 2 (skipping the header row)
    # values_only=True returns cell values instead of cell objects
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        
        # Skip totals row
        if row_dict.get("Month") == "TOTAL":
            continue
        
        # Calculate Revenue and Achievement manually if they come as formulas
        units = row_dict.get("Units Sold") or 0
        price = row_dict.get("Unit Price ($)") or 0
        target = row_dict.get("Target ($)") or 1  # avoid division by zero
        
        row_dict["Revenue ($)"] = units * price
        row_dict["Achievement (%)"] = round((units * price) / target * 100, 1)
        
        rows.append(row_dict)
    
    return headers, rows

# ─── Test block ───────────────────────────────────────────
# This only runs when you execute this file directly
# It won't run when imported by other files
if __name__ == "__main__":
    headers, rows = read_excel(filepath)
    
    print("Headers:", headers)
    print("First row:", rows[0])
    print(f"Total rows loaded: {len(rows)}")
    